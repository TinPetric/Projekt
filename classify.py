import cv2
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
import os
import openpyxl




# Give the location of the file
loc = ("oznake.xlsx")

wb = openpyxl.load_workbook(loc)
sheet = wb.active

studenti=[]

max_col = sheet.max_column
max_row = sheet.max_row

for i in range(1, max_row + 1):
    student={}
    for j in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=j)
        cell_obj = sheet.cell(row = i, column = j)
        student[cell.value]= cell_obj.value
    studenti.append(student)


list = os.listdir("segmenti2")
path = 'segmenti2\\'


test_images_brojevi=[]
test_oznake_brojevi=[]

test_images_slova=[]
test_oznake_slova=[]

for i in list:
    for student in studenti:
        if student["oznaka"] in i:
            if "jmbag" in i and "CIPsharp_20210326_111149_0008" not in i:
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append((img))
                index=int(i[-5])
                #print(i,student["jmbag"])
                broj=int(str(student["jmbag"])[index])
                test_oznake_brojevi.append(broj)
                #print(test_oznake_brojevi)
            if "bodovi" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append((img))
                index=int(i[-5])
                if student["brbod_broj"]!=None:
                    if len(str(student["brbod_broj"]))>1:
                        broj=int(str(student["brbod_broj"])[index])
                    else:
                        broj=int(student["brbod_broj"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "zadatak" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append((img))
                index=int(i[-5])
                if student["brzad"]!=None:
                    if len(str(student["brzad"]))>1:
                        broj=int(str(student["brzad"])[index])
                    else:
                        broj=int(student["brzad"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)

#print(len(test_oznake_brojevi))
#f = open("text_oznake_brojevi.txt", "a")
#f.write(str(test_oznake_brojevi))
#f.close()
#
#f = open("text_images_brojevi.txt", "a")
#f.write(str(test_images_brojevi))
#f.close()




def train():
    #mnist = tf.keras.datasets.mnist
    #(x_train, y_train), (x_test, y_test) = mnist.load_data()
    #print(x_train)
    #print(y_train)

    #x_train = tf.keras.utils.normalize(x_train, axis=1)
    #x_test = tf.keras.utils.normalize(x_test, axis=1)

#    f = open("text_oznake_brojevi.txt", "r")
#    test_oznake_brojevi=list(f.read())
#    f.close()
#
#    f = open("text_images_brojevi.txt", "r")
#    test_images_brojevi=list(f.read())
#    f.close()
#    print(test_images_brojevi)
#    print(test_oznake_brojevi)

    number=len(test_oznake_brojevi) ## ==518
    number_train=300
    x_train=np.array(test_images_brojevi[:1800])
    y_train=np.array(test_oznake_brojevi[:1800])

    x_test=np.array(test_images_brojevi[1801:2243])
    y_test=np.array(test_oznake_brojevi[1801:2243])
    #(x_train, y_train), (x_test, y_test) = mnist.load_data()

    def draw(n):
        plt.imshow(n, cmap=plt.cm.binary)
        plt.show()


    #draw(x_train[0])

    # there are two types of models
    # sequential is most common, why?

#    model = tf.keras.models.Sequential()
#
#    model.add(tf.keras.layers.Flatten(input_shape=(28,28)))
#    # reshape
#
#    model.add(tf.keras.layers.Dense(128, activation=tf.nn.relu))
#    model.add(tf.keras.layers.Dense(128, activation=tf.nn.relu))
#    model.add(tf.keras.layers.Dense(10, activation=tf.nn.softmax))
    model = tf.keras.Sequential([
        tf.keras.layers.Flatten(input_shape=(28, 28)),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(11, activation=tf.nn.softmax)
    ])

#    sample_shape = x_train[0].shape
#    print(sample_shape)
#    img_width, img_height = sample_shape[0], sample_shape[1]
#    input_shape = (img_width, img_height, 1)
#
#    x_train = x_train.reshape(len(x_train), input_shape[0], input_shape[1], input_shape[2])
#    x_test = x_test.reshape(len(x_test), input_shape[0], input_shape[1], input_shape[2])
#    #input_shape = (len(x_train),img_width, img_height, 1)
#    model = tf.keras.Sequential([
#        tf.keras.layers.Conv2D(32, (3, 3), activation='relu', kernel_initializer='he_uniform', input_shape=input_shape),
#        tf.keras.layers.MaxPooling2D((2, 2)),
#        tf.keras.layers.Flatten(),
#        tf.keras.layers.Dense(100, activation='relu', kernel_initializer='he_uniform'),
#        tf.keras.layers.Dense(10, activation=tf.nn.softmax)
#    ])


    model.compile(optimizer='adam',
                  loss='sparse_categorical_crossentropy',
                  metrics=['accuracy']
                  )
    model.fit(x_train, y_train, epochs=10)

    val_loss,val_acc = model.evaluate(x_test,y_test)
    print("loss-> ",val_loss,"\nacc-> ",val_acc)



    #draw(x_test[2])

    # saving the model
    # .h5 or .model can be used

    model.save('epic_num_reader.h5')

def clasify():

    new_model = tf.keras.models.load_model('epic_num_reader.h5')

    for i in range(2021,2244):
        predictions=new_model.predict(test_images_brojevi[i])

        print('oznaka -> ', test_oznake_brojevi[i])
        print('prediction -> ', np.argmax(predictions))

#    #path="print(np.argmax(predictions)) segmenti/0bodovi1.jpg"
#    #img = cv2.imread(path , cv2.IMREAD_GRAYSCALE)
#
#   # img=cv2.resize(img,(28,28))
#    #draw(img)
#
#    list = os.listdir("segmenti2")
#    path = 'segmenti2\\'
#
#
#    test_images=[]
#    for i in list:
#        if "jmbag" in i:
#            print(i)
#
#            img = cv2.imread(path + str(i))[:, :, 0]
#            cv2.imshow("",img)
#            img = cv2.resize(img, (28, 28))
#            img = np.invert(np.array([img]))
#            img = img / 255
#            #test_images.append((img))
#            predictions=new_model.predict(img)
#            print(np.argmax(predictions))

train()
#clasify()