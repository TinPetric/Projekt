from __future__ import absolute_import, division, print_function
from PIL import Image
import cv2
import numpy as np
#import pytesseract
import os
from PIL import Image
import tensorflow as tf

import matplotlib.pyplot as plt


import lines
import test



import os

import cv2
import tensorflow as tf
from tensorflow.keras import Model, layers
import numpy as np

import openpyxl



# Give the location of the file
loc = ("oznake.xlsx")

wb = openpyxl.load_workbook(loc)
sheet = wb.active

studenti=[]

max_col = sheet.max_column
max_row = sheet.max_row

for i in range(1, max_row + 1):
    student={}
    for j in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=j)
        cell_obj = sheet.cell(row = i, column = j)
        student[cell.value]= cell_obj.value
    studenti.append(student)





test_images_brojevi=[]
test_oznake_brojevi=[]

test_images_slova=[]
test_oznake_slova=[]


map = {'A': 0, 'B': 1, 'C': 2, 'Č': 2, 'Ć': 2, 'D': 3, 'Đ': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7,'I': 8
       ,'J': 9, 'K': 10, 'L': 11, 'M': 12,'N': 13,'O': 14, 'P': 15, 'R': 16, 'S': 17, 'Š': 17,'T': 18, 'U': 19, 'V': 20,
       'Z': 21, 'Ž': 21, ' ': 22, '-': 23}

list = os.listdir("segmenti2")
path = 'segmenti2\\'
br = 0
for i in list:
    br = 0
    for student in studenti:

        if student["oznaka"] in i:
            if "jmbag" in i: #and "CIPsharp_20210326_111149_0008" not in i:
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append(img)
                index=int(i[-5])
                #print(i,student["jmbag"])
                broj=int(str(student["jmbag"])[index])
                test_oznake_brojevi.append(broj)
                #print(test_oznake_brojevi)
            if "bodovi" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append(img)
                index=int(i[-5])
                if student["brbod_broj"]!=None:
                    if len(str(student["brbod_broj"]))>1:
                        broj=int(str(student["brbod_broj"])[index])
                    else:
                        broj=int(student["brbod_broj"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "zadatak" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append(img)
                index=int(i[-5])
                if student["brzad"]!=None:
                    if len(str(student["brzad"]))>1:
                        broj=int(str(student["brzad"])[index])
                    else:
                        broj=int(student["brzad"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "ime" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_slova.append(img)

                ime = ""
                if not student["ime1"] == None:
                    ime = student["ime1"]

                ime2 = ""
                if not student["ime2"] == None:
                    ime2 = student["ime2"]
                    ime2 = ime2 + " "

                prezime = ""
                if not student["prezime1"] == None:
                    prezime = student["prezime1"]

                prezime2 = ""
                if not student["prezime2"] == None:
                    prezime2 = student["prezime2"]
                    prezime2 = prezime2 + " "

                imeIPrezime = ime + " " + ime2 + prezime + prezime2

                imeIPrezimeChar = [char for char in imeIPrezime]
                if i[-6].isnumeric():
                    br = int(i[-6] + i[-5])
                else:
                    br = int(i[-5])

                if not br > len(imeIPrezimeChar) - 1:
                    if imeIPrezimeChar[br] == "-":
                        test_oznake_slova.append(23)
                    elif map.get(imeIPrezimeChar[br].upper()) == None:
                        test_oznake_slova.append(22)
                    else:
                        test_oznake_slova.append(map.get(imeIPrezimeChar[br].upper()))

                else:
                    test_oznake_slova.append(22)



br2 = 0
for i in test_oznake_slova:
    #print(i)
    br2 = br2 + 1

map2={}

for v in map.values():
    map2[v]=0
#print(map2)
for i in test_oznake_slova:
    #print(i)
    #print(i,map2[i])
    map2[i]+=1
print(map2)
sum=0
for i in map2:
    if i!=22:
        sum+=map2[i]
avg=sum/(len(map)-1)
#print(len(map))
brojac=0
i=0
while(1):
    if test_oznake_slova[i]==22:
        if brojac>avg:
            test_oznake_slova.pop(i)
            test_images_slova.pop(i)
            i-=1
        else: brojac+=1
    if i< (len(test_oznake_slova)-1):
        i+=1
    else:
        break

map2={}
#print(test_oznake_slova)
for v in map.values():
    map2[v]=0
#print(map2)
for i in test_oznake_slova:
    #print(i)
    #print(i,map2[i])
    map2[i]+=1
print(map2)


# MNIST dataset parameters.
num_classes = 11 # total classes (0-9 digits).

# Training parameters.
learning_rate = 0.001
training_steps = 200
batch_size = 128
display_step = 10

# Network parameters.
conv1_filters = 32 # number of filters for 1st conv layer.
conv2_filters = 64 # number of filters for 2nd conv layer.
fc1_units = 1024 # number of neurons for 1st fully-connected layer.



## Prepare MNIST data.
#from tensorflow.keras.datasets import mnist
#(x_train, y_train), (x_test, y_test) = mnist.load_data()
## Convert to float32.
#x_train, x_test = np.array(x_train, np.float32), np.array(x_test, np.float32)
## Normalize images value from [0, 255] to [0, 1].
#x_train, x_test = x_train / 255., x_test / 255.

number = len(test_oznake_brojevi)  ## ==2244
number_train = 300
x_train = np.array(test_images_brojevi[:1570])
y_train = np.array(test_oznake_brojevi[:1570])

x_test = np.array(test_images_brojevi[1571:2020])
y_test = np.array(test_oznake_brojevi[1571:2020])

# Use tf.data API to shuffle and batch data.
train_data = tf.data.Dataset.from_tensor_slices((x_train, y_train))
train_data = train_data.repeat().shuffle(5000).batch(batch_size).prefetch(1)


# Create TF Model.
class ConvNet(Model,num_classes):
    # Set layers.
    def __init__(self):
        super(ConvNet, self).__init__()
        # Convolution Layer with 32 filters and a kernel size of 5.
        self.conv1 = layers.Conv2D(32, kernel_size=5, activation=tf.nn.relu)
        # Max Pooling (down-sampling) with kernel size of 2 and strides of 2.
        self.maxpool1 = layers.MaxPool2D(2, strides=2)

        # Convolution Layer with 64 filters and a kernel size of 3.
        self.conv2 = layers.Conv2D(64, kernel_size=3, activation=tf.nn.relu)
        # Max Pooling (down-sampling) with kernel size of 2 and strides of 2.
        self.maxpool2 = layers.MaxPool2D(2, strides=2)

        # Flatten the data to a 1-D vector for the fully connected layer.
        self.flatten = layers.Flatten()

        # Fully connected layer.
        self.fc1 = layers.Dense(128)
        # Apply Dropout (if is_training is False, dropout is not applied).
        self.dropout = layers.Dropout(rate=0.5)

        # Output layer, class prediction.
        self.out = layers.Dense(num_classes)

    # Set forward pass.
    def call(self, x, is_training=False):
        x = tf.reshape(x, [-1, 28, 28, 1])
        x = self.conv1(x)
        x = self.maxpool1(x)
        x = self.conv2(x)
        x = self.maxpool2(x)
        x = self.flatten(x)
        x = self.fc1(x)
        x = self.dropout(x, training=is_training)
        x = self.out(x)
        if not is_training:
            # tf cross entropy expect logits without softmax, so only
            # apply softmax when not training.
            x = tf.nn.softmax(x)
        return x

# Build neural network model.
conv_net = ConvNet()


# Cross-Entropy Loss.
# Note that this will apply 'softmax' to the logits.
def cross_entropy_loss(x, y):
    # Convert labels to int 64 for tf cross-entropy function.
    y = tf.cast(y, tf.int64)
    # Apply softmax to logits and compute cross-entropy.
    loss = tf.nn.sparse_softmax_cross_entropy_with_logits(labels=y, logits=x)
    # Average loss across the batch.
    return tf.reduce_mean(loss)

# Accuracy metric.
def accuracy(y_pred, y_true):
    # Predicted class is the index of highest score in prediction vector (i.e. argmax).
    correct_prediction = tf.equal(tf.argmax(y_pred, 1), tf.cast(y_true, tf.int64))
    return tf.reduce_mean(tf.cast(correct_prediction, tf.float32), axis=-1)

# Stochastic gradient descent optimizer.
optimizer = tf.optimizers.Adam(learning_rate)


# Optimization process.
def run_optimization(x, y):
    # Wrap computation inside a GradientTape for automatic differentiation.
    with tf.GradientTape() as g:
        # Forward pass.
        pred = conv_net(x, is_training=True)
        # Compute loss.
        loss = cross_entropy_loss(pred, y)

    # Variables to update, i.e. trainable variables.
    trainable_variables = conv_net.trainable_variables

    # Compute gradients.
    gradients = g.gradient(loss, trainable_variables)

    # Update W and b following gradients.
    optimizer.apply_gradients(zip(gradients, trainable_variables))


# Run training for the given number of steps.
for step, (batch_x, batch_y) in enumerate(train_data.take(training_steps), 1):
    # Run the optimization to update W and b values.
    run_optimization(batch_x, batch_y)

    if step % display_step == 0:
        pred = conv_net(batch_x)
        loss = cross_entropy_loss(pred, batch_y)
        acc = accuracy(pred, batch_y)
        print("step: %i, loss: %f, accuracy: %f" % (step, loss, acc))

# Test model on validation set.
pred = conv_net(x_test)
print("Test Accuracy: %f" % accuracy(pred, y_test))

def unsharp_mask(image, kernel_size=(5, 5), sigma=5.0, amount=8.0, threshold=0):
    """Return a sharpened version of the image, using an unsharp mask."""
    blurred = cv2.GaussianBlur(image, kernel_size, sigma)
    sharpened = float(amount + 1) * image - float(amount) * blurred
    sharpened = np.maximum(sharpened, np.zeros(sharpened.shape))
    sharpened = np.minimum(sharpened, 255 * np.ones(sharpened.shape))
    sharpened = sharpened.round().astype(np.uint8)
    if threshold > 0:
        low_contrast_mask = np.absolute(image - blurred) < threshold
        np.copyto(sharpened, image, where=low_contrast_mask)


    return sharpened

map = {0:'A', 1:'B',  2:'C', 3:'D', 4:'E', 5:'F',6: 'G',7: 'H',8:'I'
       ,9:'J',10: 'K',11: 'L',12: 'M',13:'N',14:'O',15: 'P', 16:'R',17: 'S',18: 'T',19: 'U',20: 'V',
       21:'Z',22: ' ',23: '-'}

per = 25
pixelThreshold=500

roi=[[(214, 112), (1148, 152), 'text', 'ime'],
     [(216, 156), (510, 196), 'text', 'jmbag'],
     [(214, 204), (268, 240), 'text', 'zadatak'],
     [(1056, 254), (1114, 284), 'text', 'bodovi'],
     [(536, 242), (1032, 306), ' BUTTONS', 'bodovi']]



imgQ = cv2.imread('1.jpg', )
h,w,c = imgQ.shape
#imgQ = cv2.resize(imgQ,(w//3,h//3))

orb = cv2.ORB_create(1000)
kp1, des1 = orb.detectAndCompute(imgQ,None)
#impKp1 = cv2.drawKeypoints(imgQ,kp1,None)

path = 'orijentirani2'
myPicList = os.listdir(path)

#classify.train()
#new_model = tf.keras.models.load_model('epic_num_reader.h5')
for j,y in enumerate(myPicList):
    testiraj=False
    if j>=len(myPicList)-21:
        segmentipath="segmentiTest\\"
        testiraj = True

    else:
        continue
        segmentipath="segmentiTrain\\"
    if j == len(myPicList) - 21:

        import classify4

        studenti = classify4.getStudents()
        #classify4.trainNumbers()
        #modelNumbers = tf.keras.models.load_model('epic_num_reader4.h5')
        classify4.trainLetters()
        modelLetters = tf.keras.models.load_model('epic_letters_reader4.h5')

    img = cv2.imread(path + "\\" + y, cv2.IMREAD_GRAYSCALE)

    print(f'################## Extracting Data from Form {j} {y}  ##################')

    for x,r in enumerate(roi):

        imgCrop = img[r[0][1]:r[1][1], r[0][0]:r[1][0]]
        #cv2.imshow("",imgCrop)
        #cv2.waitKey(0)

        #segmentacija imena
        if x==0:
            #opet resizam sliku u ovom slucaju sliku imena i prezimena da mogu podjeliti sa hsplit
            imgCrop = cv2.resize(imgCrop, (899, 40))
            test_digits = np.hsplit(imgCrop, 31)
            result = []
            for d, z in enumerate(test_digits):
                #tu resizam pojedine slicice da budu iste velicine, objasnjeno gore
                test_digits[d] = cv2.resize(test_digits[d], (35, 40))
                #sprema u mapu segmenti da ih mozemo vidjeti lijepo
                ime = segmentipath + str(y) + "ime" + str(d) + ".jpg"

                cv2.imwrite(ime, test_digits[d])

            for d,z in enumerate(test_digits):
                ime = segmentipath + str(y) + "ime" + str(d) + ".jpg"

                cv2.imwrite(ime, test_digits[d])
                im2 = cv2.imread(ime)
                test_digits[d] = lines.remove(im2)

                test_digits[d] = cv2.resize(test_digits[d], (28, 28))

                sharpen = unsharp_mask(test_digits[d])

                _,sharpen= cv2.threshold(sharpen, 127, 255, cv2.THRESH_BINARY)

                cv2.imwrite(ime, sharpen)

                image = cv2.imread(ime)[:, :, 0]
                image = cv2.resize(image, (28, 28))
                image = np.invert(np.array([image]))
                image = image / 255

                if testiraj:
                    predictions = modelLetters.predict(image)
                    result.append(np.argmax(predictions))

            if testiraj:
                res=""
                for r in result:
                    res+=map[r]
                print(res)


        #segmentacija za jmbag
        if x==1:
            #sve isto kao i kod imena
            imgCrop = cv2.resize(imgCrop, (300, 40))
            test_digits = np.hsplit(imgCrop, 10)
            result=[]
            for d,z in enumerate(test_digits):
                test_digits[d] = cv2.resize(test_digits[d], (35, 40))
                ime = segmentipath + str(y) + "jmbag" + str(d) + ".jpg"
                #test_digits[d]=cv2.resize(test_digits[d],(28,28))
                cv2.imwrite(ime, test_digits[d])

            jmbag=[]
            for d,z in enumerate(test_digits):
                ime = segmentipath + str(y) + "jmbag" + str(d) + ".jpg"

                cv2.imwrite(ime, test_digits[d])
                im2 = cv2.imread(ime)
                test_digits[d] = lines.remove(im2)

                test_digits[d] = cv2.resize(test_digits[d], (28, 28))

                sharpen = unsharp_mask(test_digits[d])

                _,sharpen= cv2.threshold(sharpen, 127, 255, cv2.THRESH_BINARY)

                cv2.imwrite(ime, sharpen)


                image = cv2.imread(ime)[:, :, 0]
                image = cv2.resize(image, (28, 28))
                image = np.invert(np.array([image]))
                image = image / 255

                if testiraj:
                    predictions =conv_net(image)
                    result.append(np.argmax(predictions))

            if testiraj:
                res = ""
                for r in result:
                    res += str(r)
                print(res)

        #zadatak
        if x == 2:
           # cv2.imshow("1",imgCrop)
            #imgCrop = cv2.resize(imgCrop, (60, 40))
            test_digits = np.hsplit(imgCrop, 2)
            result=[]
            for d, z in enumerate(test_digits):
                test_digits[d] = cv2.resize(test_digits[d], (35, 40))
                ime = segmentipath + str(y) + "zadatak" + str(d) + ".jpg"
                cv2.imwrite(ime, test_digits[d])

            for d,z in enumerate(test_digits):
                ime = segmentipath + str(y) + "zadatak" + str(d) + ".jpg"

                cv2.imwrite(ime, test_digits[d])
                im2 = cv2.imread(ime)
                test_digits[d] = lines.remove(im2)

                test_digits[d] = cv2.resize(test_digits[d], (28, 28))
                sharpen = unsharp_mask(test_digits[d])
                _,sharpen= cv2.threshold(sharpen, 127, 255, cv2.THRESH_BINARY)

                cv2.imwrite(ime, sharpen)
                image = cv2.imread(ime)[:, :, 0]
                image = cv2.resize(image, (28, 28))
                image = np.invert(np.array([image]))
                image = image / 255

                if testiraj:
                    predictions =conv_net(image)
                    result.append(np.argmax(predictions))

            if testiraj:
                res = ""
                for r in result:
                    res += str(r)
                print(res)

        #bodovi
        if x==3:
            imgCrop = cv2.resize(imgCrop, (60, 40))
            test_digits = np.hsplit(imgCrop, 2)
            for d, z in enumerate(test_digits):
                test_digits[d] = cv2.resize(test_digits[d], (35, 40))
                ime = segmentipath + str(y) + "bodovi" + str(d) + ".jpg"
                cv2.imwrite(ime, test_digits[d])

            result=[]
            for d,z in enumerate(test_digits):
                ime = segmentipath + str(y) + "bodovi" + str(d) + ".jpg"

                cv2.imwrite(ime, test_digits[d])
                im2 = cv2.imread(ime)
                test_digits[d] = lines.remove(im2)

                test_digits[d] = cv2.resize(test_digits[d], (28, 28))
               # crop_img = test_digits[d][2:24, 3:25]
                #crop_img = cv2.resize(crop_img, (28, 28))
                sharpen = unsharp_mask(test_digits[d])

                _,sharpen= cv2.threshold(sharpen, 127, 255, cv2.THRESH_BINARY)

                cv2.imwrite(ime, sharpen)
                image = cv2.imread(ime)[:, :, 0]
                image = cv2.resize(image, (28, 28))
                image = np.invert(np.array([image]))
                image = image / 255

                if testiraj:
                    predictions =conv_net(image)
                    result.append(np.argmax(predictions))

            if testiraj:
                res = ""
                for r in result:
                    res += str(r)
                print(res)



        if x==4:
            ime = segmentipath + str(y) + "Bodovi.jpg"
            cv2.imwrite(ime, imgCrop)
            image=cv2.imread(ime)
            res=test.readButton(image)
            print(res)

        cv2.waitKey(0)



cv2.waitKey(0)

