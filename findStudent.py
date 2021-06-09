import difflib

import Levenshtein as lev
#import classify4


#students=classify4.getStudents()
import os
import openpyxl




# Give the location of the file
loc = ("oznake.xlsx")

wb = openpyxl.load_workbook(loc)
sheet = wb.active

studenti=[]

max_col = sheet.max_column
max_row = sheet.max_row

for i in range(1, max_row + 1):
    student={}
    for j in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=j)
        cell_obj = sheet.cell(row = i, column = j)
        student[cell.value]= cell_obj.value
    studenti.append(student)

def getStudents():
    result=[]

    for student in studenti:
        st={}
        ime = ""
        ime+=student["ime1"]
        if student["ime2"]!=None:
            ime+=" "+student["ime2"]
        if student["prezime1"] != None:
            ime += " " + student["prezime1"]
        if student["prezime2"] != None:
            ime += " " + student["prezime2"]
        st["ime"]=ime
        st["jmbag"]=student["jmbag"]
        st["brzad"] = student["brzad"]

        if student["brbod_broj"]:
            st["brbod_broj"] = student["brbod_broj"]
        else:
            st["brbod_broj"]=" "
        if student["brbod_broj"]:
            st["brbod"] = student["brbod"]
        else:
            st["brbod"]=" "

        result.append(st)
    return result

def find1(imeprezime,jmbag):
    global students
    #print(students)
    i=0
    while(1):
        if imeprezime[i]==" ":
            if imeprezime[i+1]==" ":
                imeprezime=imeprezime[:i].strip()
                break
        i+=1
    moguci = []
    for i in students:
        if len(i["ime"]) == len(imeprezime):
            if " " in imeprezime:
                ime1 = imeprezime.split()[0]
                prezime1 = imeprezime.split()[1]
                ime2 = i["ime"].split()[0]
                prezime2 = i["ime"].split()[1]
                if len(ime1) == len(ime2) and len(prezime1) == len(prezime2):
                    moguci.append(i)

            else:
                moguci.append(i)
    ratios=[]
    #print(moguci)

    for s in moguci:
        Ratio = lev.ratio(imeprezime, s["ime"])

        ratios.append(Ratio)
    #print(ratios)

    maxRat=max(ratios)

    moguci3 = []
    for i in range(0,len(ratios)):
        if ratios[i] == maxRat:
            moguci3.append(moguci[i])
    #print(moguci3)

    ratios = []

    jmbagovi=[i["jmbag"] for i in students]
    for s in moguci3:

        Ratio = lev.ratio(jmbag, s["jmbag"])
        # print(Ratio)
        ratios.append(Ratio)

    #print(ratios)

    maxRat = max(ratios)

    moguci5 = []
    for i in range(0, len(ratios)):
        if ratios[i] == maxRat:
            moguci5.append(moguci3[i])
    #print(moguci5)
    return moguci5

def find2(imeprezime,jmbag):
    global students
    #print(students)
    i=0
    while(1):
        if imeprezime[i]==" ":
            if imeprezime[i+1]==" ":
                imeprezime=imeprezime[:i].strip()
                break
        i+=1
    moguci = []
    for i in students:
        if len(i["ime"]) == len(imeprezime):
            if " " in imeprezime:
                ime1 = imeprezime.split()[0]
                prezime1 = imeprezime.split()[1]
                ime2 = i["ime"].split()[0]
                prezime2 = i["ime"].split()[1]
                if len(ime1) == len(ime2) and len(prezime1) == len(prezime2):
                    moguci.append(i)

            else:
                moguci.append(i)
    distances=[]
    #print(moguci)

    for s in moguci:
        Distance = lev.distance(imeprezime, s["ime"])
        #print(Distance)
        distances.append(Distance)


    #print(distances)
    minDis=min(distances)
    moguci2=[]
    for i in range(0,len(distances)):
        if distances[i]==minDis:
            moguci2.append(moguci[i])
    #print(moguci2)

    distances = []
    jmbagovi=[i["jmbag"] for i in students]
    for s in moguci2:
        Distance = lev.distance(jmbag, s["jmbag"])
        # print(Distance)
        distances.append(Distance)

    #print(distances)

    minDis = min(distances)
    moguci4 = []
    for i in range(0, len(distances)):
        if distances[i] == minDis:
            moguci4.append(moguci2[i])
    #print(moguci4)
    return moguci4


students=getStudents()
#find("IOZE KDLUP                     ","6036181392")
#find("ELRLKA LUOIA              ","0036301015")
#find1("EMARAM BADDM                   ","0036426914")
#find("NARIJAO SVCAK                  ","6463133110")

#find2("IOZE KDLUP                     ","6036181392")
#find2("ELRLKA LUOIA              ","0036301015")
#find2("EMARAM BADDM                   ","0036426914")
#find2("NARIJAO SVCAK                  ","6463133110")
#
#find1("IOZE KDLUP                     ","6036181392")
#find1("ELRLKA LUOIA              ","0036301015")
#find1("EMARAM BADDM                   ","0036426914")
#find1("NARIJAO SVCAK                  ","6463133110")

#Ratio = lev.ratio("6036181392", '0036518943')
#print(Ratio)
#Distance = lev.distance("6036181392", '0036518943')
#print(Distance)
#Ratio = lev.ratio("6036181392", '2456781385')
#print(Ratio)
#
#Distance = lev.distance("6036181392", '2456781385')
#print(Distance)