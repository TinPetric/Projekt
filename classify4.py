import cv2
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
import os
import openpyxl
import math




# Give the location of the file
loc = ("oznake.xlsx")

wb = openpyxl.load_workbook(loc)
sheet = wb.active

studenti=[]

max_col = sheet.max_column
max_row = sheet.max_row

for i in range(1, max_row + 1):
    student={}
    for j in range(1, max_col + 1):
        cellname = sheet.cell(row=1, column=j)
        cell_obj = sheet.cell(row = i, column = j)
        student[cellname.value]= cell_obj.value
    studenti.append(student)





test_images_brojevi=[]
test_oznake_brojevi=[]

test_images_slova=[]
test_oznake_slova=[]


map = {'A': 0, 'B': 1, 'C': 2, 'Č': 2, 'Ć': 2, 'D': 3, 'Đ': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7,'I': 8
       ,'J': 9, 'K': 10, 'L': 11, 'M': 12,'N': 13,'O': 14, 'P': 15, 'R': 16, 'S': 17, 'Š': 17,'T': 18, 'U': 19, 'V': 20,
       'Z': 21, 'Ž': 21, ' ': 22, '-': 23}

list = os.listdir("segmentiTrain")
path = 'segmentiTrain\\'
br = 0
for i in list:
    br = 0
    for student in studenti:

        if student["oznaka"] in i:
            if "jmbag" in i: #and "CIPsharp_20210326_111149_0008" not in i:
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append(img)
                index=int(i[-5])
                #print(i,student["jmbag"])
                broj=int(str(student["jmbag"])[index])
                test_oznake_brojevi.append(broj)
                #print(test_oznake_brojevi)
            if "bodovi" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append(img)
                index=int(i[-5])
                if student["brbod_broj"]!=None:
                    if len(str(student["brbod_broj"]))>1:
                        broj=int(str(student["brbod_broj"])[index])
                    else:
                        broj=int(student["brbod_broj"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "zadatak" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append(img)
                index=int(i[-5])
                if student["brzad"]!=None:
                    if len(str(student["brzad"]))>1:
                        broj=int(str(student["brzad"])[index])
                    else:
                        broj=int(student["brzad"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "ime" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_slova.append(img)

                ime = ""
                if not student["ime1"] == None:
                    ime = student["ime1"]

                ime2 = ""
                if not student["ime2"] == None:
                    ime2 = student["ime2"]
                    ime2 = ime2 + " "

                prezime = ""
                if not student["prezime1"] == None:
                    prezime = student["prezime1"]

                prezime2 = ""
                if not student["prezime2"] == None:
                    prezime2 = student["prezime2"]
                    prezime2 = prezime2 + " "

                imeIPrezime = ime + " " + ime2 + prezime + prezime2

                imeIPrezimeChar = [char for char in imeIPrezime]
                if i[-6].isnumeric():
                    br = int(i[-6] + i[-5])
                else:
                    br = int(i[-5])

                if not br > len(imeIPrezimeChar) - 1:
                    if imeIPrezimeChar[br] == "-":
                        test_oznake_slova.append(23)
                    elif map.get(imeIPrezimeChar[br].upper()) == None:
                        test_oznake_slova.append(22)
                    else:
                        test_oznake_slova.append(map.get(imeIPrezimeChar[br].upper()))

                else:
                    test_oznake_slova.append(22)



br2 = 0
for i in test_oznake_slova:
    #print(i)
    br2 = br2 + 1

map2={}

for v in map.values():
    map2[v]=0
#print(map2)
for i in test_oznake_slova:
    #print(i)
    #print(i,map2[i])
    map2[i]+=1
print(map2)
sum=0
for i in map2:
    if i!=22:
        sum+=map2[i]
avg=sum/(len(map)-1)
#print(len(map))
brojac=0
i=0
while(1):
    if test_oznake_slova[i]==22:
        if brojac>avg:
            test_oznake_slova.pop(i)
            test_images_slova.pop(i)
            i-=1
        else: brojac+=1
    if i< (len(test_oznake_slova)-1):
        i+=1
    else:
        break

map2={}
#print(test_oznake_slova)
for v in map.values():
    map2[v]=0
#print(map2)
for i in test_oznake_slova:
    #print(i)
    #print(i,map2[i])
    map2[i]+=1
print(map2)

number=len(test_oznake_slova) #==5768 5766
print(number,len(test_images_slova))

def getStudents():
    result=[]

    for student in studenti:
        st={}
        ime = ""
        ime+=student["ime1"]
        if student["ime2"]!=None:
            ime+=" "+student["ime2"]
        if student["prezime1"] != None:
            ime += " " + student["prezime1"]
        if student["prezime2"] != None:
            ime += " " + student["prezime2"]
        st["ime"]=ime
        st["jmbag"]=student["jmbag"]
        st["brzad"] = student["brzad"]
        if student["brbod_broj"]:
            st["brbod_broj"] = student["brbod_broj"]
        else:
            st["brbod_broj"]=" "
        result.append(st)
    return result


def trainNumbers():


    number=math.floor(len(test_oznake_brojevi)*0.9) ## ==2244
    number_train=300
    x_train = np.array(test_images_brojevi[:number])
    y_train = np.array(test_oznake_brojevi[:number])

    x_test = np.array(test_images_brojevi[number + 1:len(test_oznake_brojevi) - 1])
    y_test = np.array(test_oznake_brojevi[number + 1:len(test_oznake_brojevi) - 1])


    def draw(n):
        plt.imshow(n, cmap=plt.cm.binary)
        plt.show()


    model = tf.keras.Sequential([
        tf.keras.layers.Flatten(input_shape=(28, 28)),
        #tf.keras.layers.Dense(512, activation=tf.nn.relu),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(11, activation=tf.nn.softmax)
    ])



    model.compile(optimizer='adam',
                  loss='sparse_categorical_crossentropy',
                  metrics=['accuracy']
                  )
    model.fit(x_train, y_train, epochs=10)

    val_loss,val_acc = model.evaluate(x_test,y_test)
    print("loss-> ",val_loss,"\nacc-> ",val_acc)


    model.save('epic_num_reader4.h5')

def trainLetters():

#{0: 297, 1: 35, 2: 10, 3: 19, 4: 119, 5: 55, 6: 8, 7: 102, 8: 10, 9: 22, 10: 10,
# 11: 271, 12: 57, 13: 101, 14: 87, 15: 105, 16: 135, 17: 165, 18: 49, 19: 184, 20: 39, 21:20,
# 22: 73, 23: 66, 24: 119, 25: 20, 26: 7, 27: 3579, 28: 4}

    number=math.floor(len(test_oznake_slova)*0.9)

    x_train=np.array(test_images_slova[:number])
    y_train=np.array(test_oznake_slova[:number])

    x_test=np.array(test_images_slova[number+1:len(test_oznake_slova)-1])
    y_test=np.array(test_oznake_slova[number+1:len(test_oznake_slova)-1])
    #(x_train, y_train), (x_test, y_test) = mnist.load_data()

    def draw(n):
        plt.imshow(n, cmap=plt.cm.binary)
        plt.show()


    model = tf.keras.Sequential([
        tf.keras.layers.Flatten(input_shape=(28, 28)),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(24, activation=tf.nn.softmax)
    ])


    model.compile(optimizer='adam',
                  loss='sparse_categorical_crossentropy',
                  metrics=['accuracy']
                  )
    model.fit(x_train, y_train, epochs=10)

    val_loss,val_acc = model.evaluate(x_test,y_test)
    print("loss-> ",val_loss,"\nacc-> ",val_acc)


    model.save('epic_letters_reader4.h5')

def classifyNumbers():

    new_model = tf.keras.models.load_model('epic_num_reader.h5')

    for i in range(2021,2244):
        predictions=new_model.predict(test_images_brojevi[i])

        print('oznaka -> ', test_oznake_brojevi[i])
        print('prediction -> ', np.argmax(predictions))

def classifyLetters():

    new_model = tf.keras.models.load_model('epic_letters_reader.h5')

    for i in range(2021,2244):
        predictions=new_model.predict(test_images_slova[i])

        print('oznaka -> ', test_oznake_slova[i])
        print('prediction -> ', np.argmax(predictions))


#trainNumbers()
#classifyNumbers()
#trainLetters()
#classifyLetters()