import cv2
import numpy as np
import matplotlib.pyplot as plt
import tensorflow as tf
import os
import openpyxl




# Give the location of the file
loc = ("oznake.xlsx")

wb = openpyxl.load_workbook(loc)
sheet = wb.active

studenti=[]

max_col = sheet.max_column
max_row = sheet.max_row

for i in range(1, max_row + 1):
    student={}
    for j in range(1, max_col + 1):
        cell = sheet.cell(row=1, column=j)
        cell_obj = sheet.cell(row = i, column = j)
        student[cell.value]= cell_obj.value
    studenti.append(student)


list = os.listdir("segmenti2")
path = 'segmenti2\\'


test_images_brojevi=[]
test_oznake_brojevi=[]

test_images_slova=[]
test_oznake_slova=[]


map = {'A': 0, 'B': 1, 'C': 2, 'Č': 3, 'Ć': 4, 'D': 5, 'Đ': 6, 'E': 7, 'F': 8, 'G': 9, 'H': 10,'I': 11
       ,'J': 12, 'K': 13, 'L': 14, 'M': 15,'N': 16,'O': 17, 'P': 18, 'R': 19, 'S': 20, 'Š': 21,'T': 22, 'U': 23, 'V': 24,
       'Z': 25, 'Ž': 26, ' ': 27}
br = 0
for i in list:
    if br > 30:
        br = 0
    for student in studenti:

        if student["oznaka"] in i:
            if "jmbag" in i and "CIPsharp_20210326_111149_0008" not in i:
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append((img))
                index=int(i[-5])
                #print(i,student["jmbag"])
                broj=int(str(student["jmbag"])[index])
                test_oznake_brojevi.append(broj)
                #print(test_oznake_brojevi)
            if "bodovi" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append((img))
                index=int(i[-5])
                if student["brbod_broj"]!=None:
                    if len(str(student["brbod_broj"]))>1:
                        broj=int(str(student["brbod_broj"])[index])
                    else:
                        broj=int(student["brbod_broj"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "zadatak" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_brojevi.append((img))
                index=int(i[-5])
                if student["brzad"]!=None:
                    if len(str(student["brzad"]))>1:
                        broj=int(str(student["brzad"])[index])
                    else:
                        broj=int(student["brzad"])
                else:
                    broj=10
                test_oznake_brojevi.append(broj)
            if "ime" in i :
                img = cv2.imread(path + str(i))[:, :, 0]
                img = cv2.resize(img, (28, 28))
                img = np.invert(np.array([img]))
                img = img / 255
                test_images_slova.append((img))


                ime = student["ime1"]
                if not student["ime2"] == None:
                    ime2 = student["ime2"]
                    imeChar2 = [char for char in ime2]
                    ime2Duljina = len(imeChar2)
                if not student["prezime2"] == None:
                    prezime2 = student["prezime2"]
                    prezimeChar2 = [char for char in prezime2]
                    prezime2Duljina = len(prezimeChar2)
                prezime = student["prezime1"]

                imeChar = [char for char in ime]
                if not student["prezime1"] == None:
                    prezimeChar = [char for char in prezime]

                imeDuljina = len(imeChar)
                prezimeDuljina = len(prezimeChar)
                pomocni = 0
                pomocni2 = 0
                if br < imeDuljina:
                    slovo = imeChar[br]
                    brojSlova = map.get(slovo.upper())
                    if brojSlova!=None:
                        test_oznake_slova.append(brojSlova)
                    else:
                        test_oznake_slova.append(27)
                if br == imeDuljina:
                    test_oznake_slova.append(27)
                if br > imeDuljina:
                    if not student["ime2"] == None:
                        if br - imeDuljina  < ime2Duljina:
                            slovo = imeChar2[br - imeDuljina - 1]
                            brojSlova = map.get(slovo.upper())
                            if brojSlova != None:
                                test_oznake_slova.append(brojSlova)
                            else:
                                test_oznake_slova.append(27)
                        if br - imeDuljina == ime2Duljina:
                            test_oznake_slova.append(27)
                        pomocni = ime2Duljina
                        if br > imeDuljina + pomocni:
                            if br - imeDuljina - pomocni -1 <  prezimeDuljina:
                                slovo = prezimeChar[br - imeDuljina - pomocni-1]
                                brojSlova = map.get(slovo.upper())
                                if brojSlova != None:
                                    test_oznake_slova.append(brojSlova)
                                else:
                                    test_oznake_slova.append(27)
                        if br > imeDuljina + pomocni + prezimeDuljina:

                            if not student["prezime2"] == None:
                                if br - imeDuljina - pomocni == prezimeDuljina:
                                    test_oznake_slova.append(27)
                                if br - imeDuljina - pomocni - prezimeDuljina < prezime2Duljina:
                                    slovo = prezimeChar2[br - imeDuljina - pomocni-1 - prezimeDuljina]
                                    brojSlova = map.get(slovo.upper())
                                    if brojSlova != None:
                                        test_oznake_slova.append(brojSlova)
                                    else:
                                        test_oznake_slova.append(27)
                                pomocni2 = prezime2Duljina
                    if br > imeDuljina + pomocni + prezimeDuljina + pomocni2:
                        test_oznake_slova.append(27)

                br  = br + 1


#print(len(test_oznake_slova))
#br2 = 0
#for i in test_oznake_slova:
#    if  br2 < 31:
#        print(i)
#    br2 = br2 + 1
#print(test_images_slova)
#print(test_oznake_slova)
map2={}
print(test_oznake_slova)
for v in map.values():
    map2[v]=0
print(map2)
for i in test_oznake_slova:
    print(i)
    print(i,map2[i])
    map2[i]+=1
print(map2)



def trainNumbers():


    number=len(test_oznake_brojevi) ## ==2244
    number_train=300
    x_train=np.array(test_images_brojevi[:1570])
    y_train=np.array(test_oznake_brojevi[:1570])

    x_test=np.array(test_images_brojevi[1571:2020])
    y_test=np.array(test_oznake_brojevi[1571:2020])


    def draw(n):
        plt.imshow(n, cmap=plt.cm.binary)
        plt.show()


    model = tf.keras.Sequential([
        tf.keras.layers.Flatten(input_shape=(28, 28)),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(11, activation=tf.nn.softmax)
    ])



    model.compile(optimizer='adam',
                  loss='sparse_categorical_crossentropy',
                  metrics=['accuracy']
                  )
    model.fit(x_train, y_train, epochs=10)

    val_loss,val_acc = model.evaluate(x_test,y_test)
    print("loss-> ",val_loss,"\nacc-> ",val_acc)


    model.save('epic_num_reader2.h5')

def trainLetters():


    number=len(test_oznake_slova) #==4987,4991
    print(number,len(test_images_slova))
    x_train=np.array(test_images_slova[:1570])
    y_train=np.array(test_oznake_slova[:1570])

    x_test=np.array(test_images_slova[1571:2020])
    y_test=np.array(test_oznake_slova[1571:2020])
    #(x_train, y_train), (x_test, y_test) = mnist.load_data()

    def draw(n):
        plt.imshow(n, cmap=plt.cm.binary)
        plt.show()


    model = tf.keras.Sequential([
        tf.keras.layers.Flatten(input_shape=(28, 28)),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(1024, activation=tf.nn.relu),
        tf.keras.layers.Dense(28, activation=tf.nn.softmax)
    ])


    model.compile(optimizer='adam',
                  loss='sparse_categorical_crossentropy',
                  metrics=['accuracy']
                  )
    model.fit(x_train, y_train, epochs=10)

    val_loss,val_acc = model.evaluate(x_test,y_test)
    print("loss-> ",val_loss,"\nacc-> ",val_acc)


    model.save('epic_letters_reader2.h5')

def classifyNumbers():

    new_model = tf.keras.models.load_model('epic_num_reader.h5')

    for i in range(2021,2244):
        predictions=new_model.predict(test_images_brojevi[i])

        print('oznaka -> ', test_oznake_brojevi[i])
        print('prediction -> ', np.argmax(predictions))

def classifyLetters():

    new_model = tf.keras.models.load_model('epic_letters_reader.h5')

    for i in range(2021,2244):
        predictions=new_model.predict(test_images_slova[i])

        print('oznaka -> ', test_oznake_slova[i])
        print('prediction -> ', np.argmax(predictions))


#trainNumbers()
#classifyNumbers()
trainLetters()
classifyLetters()